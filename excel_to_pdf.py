import pandas as pd
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from datetime import datetime
import os


class ExcelToPDFConverter:
    def __init__(self, excel_file, output_pdf=None):
        self.excel_file = excel_file
        self.output_pdf = output_pdf or f"{os.path.splitext(excel_file)[0]}_extracted.pdf"
        self.wb = load_workbook(excel_file, data_only=True)
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
        
    def _setup_custom_styles(self):
        """Setup custom paragraph styles"""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        self.styles.add(ParagraphStyle(
            name='SheetTitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2e5c8a'),
            spaceAfter=10,
            spaceBefore=10,
            fontName='Helvetica-Bold'
        ))
        
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading3'],
            fontSize=11,
            textColor=colors.HexColor('#444444'),
            spaceAfter=6,
            fontName='Helvetica-Bold'
        ))
    
    def _clean_dataframe(self, df):
        """Clean and prepare dataframe for PDF"""
        # Replace NaN with empty string
        df = df.fillna('')
        
        # Convert all values to string
        df = df.astype(str)
        
        # Remove rows that are completely empty
        df = df[~(df == '').all(axis=1)]
        
        return df
    
    def _create_table_from_dataframe(self, df, sheet_name):
        """Create a formatted table from dataframe"""
        # Clean the dataframe
        df_clean = self._clean_dataframe(df)
        
        if df_clean.empty:
            return None
        
        # Convert dataframe to list of lists for ReportLab
        data = [df_clean.columns.tolist()] + df_clean.values.tolist()
        
        # Create table
        table = Table(data, repeatRows=1)
        
        # Define table style based on sheet type
        style = self._get_table_style(sheet_name)
        table.setStyle(style)
        
        return table
    
    def _get_table_style(self, sheet_name):
        """Get appropriate table style based on sheet name"""
        base_style = [
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1f4788')),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ]
        
        return TableStyle(base_style)
    
    def _extract_data_sheet_info(self):
        """Extract metadata from Data Sheet"""
        try:
            df = pd.read_excel(self.excel_file, sheet_name='Data Sheet', header=None)
            
            info = {
                'company_name': 'APOLLO TYRES LTD',
                'version': '',
                'extraction_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # Try to extract version
            for idx, row in df.iterrows():
                if 'CURRENT VERSION' in str(row[0]):
                    info['version'] = str(row[1]) if pd.notna(row[1]) else ''
                    break
            
            return info
        except Exception as e:
            print(f"Warning: Could not extract metadata: {e}")
            return {
                'company_name': 'APOLLO TYRES LTD',
                'version': 'N/A',
                'extraction_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
    
    def convert(self):
        """Main conversion method"""
        print(f"Starting conversion of {self.excel_file}...")
        
        # Create PDF document
        doc = SimpleDocTemplate(
            self.output_pdf,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # Container for PDF elements
        elements = []
        
        # Extract metadata
        metadata = self._extract_data_sheet_info()
        
        # Add title page
        elements.append(Paragraph(
            f"Financial Data Extract: {metadata['company_name']}", 
            self.styles['CustomTitle']
        ))
        elements.append(Spacer(1, 0.2*inch))
        
        # Add metadata
        metadata_text = f"""
        <b>Extraction Date:</b> {metadata['extraction_date']}<br/>
        <b>Source File:</b> {self.excel_file}<br/>
        <b>Version:</b> {metadata['version']}<br/>
        """
        elements.append(Paragraph(metadata_text, self.styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Process each sheet (except Customization and Data Sheet)
        sheets_to_process = [
            'Profit & Loss',
            'Quarters', 
            'Balance Sheet',
            'Cash Flow'
        ]
        
        for sheet_name in sheets_to_process:
            if sheet_name not in self.wb.sheetnames:
                continue
                
            print(f"Processing sheet: {sheet_name}")
            
            # Add sheet title
            elements.append(PageBreak())
            elements.append(Paragraph(sheet_name, self.styles['SheetTitle']))
            elements.append(Spacer(1, 0.1*inch))
            
            # Read the sheet with proper handling
            df = self._read_sheet_intelligently(sheet_name)
            
            if df is not None and not df.empty:
                # Create table
                table = self._create_table_from_dataframe(df, sheet_name)
                
                if table:
                    elements.append(table)
                    elements.append(Spacer(1, 0.2*inch))
                else:
                    elements.append(Paragraph(
                        f"No data available for {sheet_name}", 
                        self.styles['Normal']
                    ))
            else:
                elements.append(Paragraph(
                    f"No data available for {sheet_name}", 
                    self.styles['Normal']
                ))
        
        # Build PDF
        print("Building PDF...")
        doc.build(elements)
        print(f"✅ PDF created successfully: {self.output_pdf}")
        
        self.wb.close()
        return self.output_pdf
    
    def _read_sheet_intelligently(self, sheet_name):
        """Read sheet with intelligent header detection"""
        try:
            # Read the raw data
            df_raw = pd.read_excel(self.excel_file, sheet_name=sheet_name, header=None)
            
            # Find the header row (look for 'Narration' or similar keywords)
            header_row = None
            for idx, row in df_raw.iterrows():
                if 'Narration' in str(row.values):
                    header_row = idx
                    break
            
            if header_row is not None:
                # Re-read with proper header
                df = pd.read_excel(
                    self.excel_file, 
                    sheet_name=sheet_name, 
                    header=header_row
                )
                
                # Clean column names
                df.columns = [str(col).strip() if pd.notna(col) else f'Col_{i}' 
                             for i, col in enumerate(df.columns)]
                
                # Remove rows before the header
                df = df[df.index > header_row]
                
                # Remove completely empty rows
                df = df.dropna(how='all')
                
                return df
            else:
                # If no header found, use the raw data
                return df_raw.dropna(how='all')
                
        except Exception as e:
            print(f"Error reading sheet {sheet_name}: {e}")
            return None


def main():
    """Main execution function"""
    excel_file = "Apollo Tyres.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Error: File '{excel_file}' not found!")
        return
    
    # Create converter instance
    converter = ExcelToPDFConverter(excel_file)
    
    # Convert to PDF
    output_file = converter.convert()
    
    print(f"\n{'='*80}")
    print(f"✅ Conversion Complete!")
    print(f"📄 Output PDF: {output_file}")
    print(f"{'='*80}")


if __name__ == "__main__":
    main()