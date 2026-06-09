from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from datetime import datetime
import os
import re


class DataSheetToPDFExtractor:
    def __init__(self, excel_file, output_pdf=None):
        self.excel_file = excel_file
        self.output_pdf = output_pdf or f"{os.path.splitext(excel_file)[0]}_extracted_v2.pdf"
        self.wb = load_workbook(excel_file, data_only=True)
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
        
    def _setup_custom_styles(self):
        """Setup custom paragraph styles"""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=12,
            alignment=1
        ))
        
        self.styles.add(ParagraphStyle(
            name='SectionTitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=6,
            spaceBefore=12
        ))
    
    def _format_date(self, value):
        """Format date value to Mar-17 format"""
        if not value:
            return ''
        
        value_str = str(value).strip()
        
        # If it's already formatted like "Mar-17", keep it
        if re.match(r'[A-Za-z]{3}-\d{2}', value_str):
            return value_str
        
        # If it's a full date string like "2017-03-31 00:00:00"
        if re.match(r'\d{4}-\d{2}-\d{2}', value_str):
            try:
                # Extract year and month
                parts = value_str.split()[0].split('-')
                year = parts[0]
                month = parts[1]
                
                # Convert to Mar-17 format
                month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                month_abbr = month_names[int(month)]
                year_short = year[-2:]
                
                return f"{month_abbr}-{year_short}"
            except:
                pass
        
        # If it's a datetime object
        try:
            if isinstance(value, datetime):
                return value.strftime('%b-%y')
        except:
            pass
        
        return value_str
    
    def _clean_cell_value(self, value):
        """Clean cell value"""
        if value is None:
            return ''
        
        value_str = str(value).strip()
        
        # Remove URLs and screener references
        if 'screener.in' in value_str.lower() or 'http' in value_str.lower():
            return ''
        
        return value_str
    
    def _extract_data_sheet(self):
        """Extract all data from Data Sheet"""
        try:
            ws = self.wb['Data Sheet']
            
            sections = {
                'company_info': [],  # Combined section for company info and metadata
                'profit_loss': [],
                'quarters': [],
                'balance_sheet': [],
                'cash_flow': [],
                'price_market': []  # Combined section
            }
            
            current_section = None
            
            for row in ws.iter_rows(values_only=True):
                # Convert to list and clean values
                row_list = [self._clean_cell_value(cell) for cell in row]
                
                # Skip completely empty rows
                if not any(cell for cell in row_list):
                    continue
                
                first_cell = row_list[0].strip().upper() if row_list[0] else ''
                
                # Identify sections - combine COMPANY NAME and META into company_info
                if 'COMPANY NAME' in first_cell or first_cell == 'META':
                    current_section = 'company_info'
                    if 'COMPANY NAME' in first_cell:
                        sections[current_section].append(row_list)
                    continue
                elif first_cell == 'PROFIT & LOSS':
                    current_section = 'profit_loss'
                    continue
                elif first_cell == 'QUARTERS':
                    current_section = 'quarters'
                    continue
                elif first_cell == 'BALANCE SHEET':
                    current_section = 'balance_sheet'
                    continue
                elif first_cell == 'CASH FLOW:':
                    current_section = 'cash_flow'
                    continue
                elif first_cell == 'PRICE:' or first_cell == 'DERIVED:':
                    # Combine PRICE and DERIVED into price_market section
                    current_section = 'price_market'
                    if first_cell == 'PRICE:':
                        sections[current_section].append(row_list)
                    continue
                elif 'REPORT DATE' in first_cell:
                    # Format report dates properly
                    formatted_dates = [self._format_date(cell) for cell in row_list[1:] if cell]
                    if current_section:
                        sections[current_section].append(['Report Date'] + formatted_dates)
                elif current_section and row_list[0]:
                    # Skip excluded rows
                    if any(x in first_cell for x in ['LATEST VERSION', 'CURRENT VERSION', 'PLEASE DO NOT']):
                        continue
                    sections[current_section].append(row_list)
            
            return sections
            
        except Exception as e:
            print(f"Error extracting data: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
    def _create_table(self, data, section_name=''):
        """Create a formatted table for the PDF"""
        if not data:
            return None
        
        # Filter out empty rows and prepare table data
        table_data = []
        max_cols = 0
        
        for row in data:
            # Find the last non-empty cell
            last_idx = len(row) - 1
            while last_idx > 0 and not str(row[last_idx]).strip():
                last_idx -= 1
            
            # Keep row up to last non-empty cell
            filtered_row = row[:last_idx + 1]
            
            if filtered_row and any(str(cell).strip() for cell in filtered_row):
                table_data.append(filtered_row)
                max_cols = max(max_cols, len(filtered_row))
        
        if not table_data:
            return None
        
        # Pad all rows to have the same number of columns
        for row in table_data:
            while len(row) < max_cols:
                row.append('')
        
        # Determine if first row is a header (contains "Report Date" or is all text)
        is_header_row = 'Report Date' in str(table_data[0][0]) if table_data else False
        
        # Calculate column widths
        if max_cols <= 3:
            col_widths = [3*inch] + [2*inch] * (max_cols - 1)
        else:
            col_widths = [2.5*inch] + [0.85*inch] * (max_cols - 1)
        
        table = Table(table_data, colWidths=col_widths)
        
        # Style the table
        style_commands = [
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # First column (metric names) - bold and left aligned
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            
            # Data columns - right aligned
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (1, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]
        
        # If first row is header
        if is_header_row:
            style_commands.extend([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2c3e50')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.lightgrey]),
            ])
        else:
            style_commands.extend([
                ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.beige, colors.lightgrey]),
            ])
        
        table.setStyle(TableStyle(style_commands))
        return table
    
    def convert(self):
        """Main conversion method"""
        print(f"\n{'='*80}")
        print(f"Extracting data from Data Sheet to PDF (v2)...")
        print(f"{'='*80}\n")
        
        sections = self._extract_data_sheet()
        
        if not sections:
            print("❌ No data extracted!")
            return None
        
        # Extract company name
        company_name = "APOLLO HOSPITALS ENTERPRISE LTD"
        if sections.get('company_info'):
            for row in sections['company_info']:
                if row[0] and 'COMPANY NAME' in str(row[0]).upper():
                    company_name = row[1] if len(row) > 1 and row[1] else company_name
                    break
        
        doc = SimpleDocTemplate(
            self.output_pdf,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        elements = []
        
        # Title
        elements.append(Paragraph(
            f"{company_name} - Financial Data Extract", 
            self.styles['CustomTitle']
        ))
        elements.append(Spacer(1, 0.2*inch))
        
        # Metadata
        current_date = datetime.now().strftime('%d %B %Y, %I:%M %p')
        metadata_text = f"""
        <b>Extraction Date:</b> {current_date}<br/>
        <b>Source File:</b> {self.excel_file}<br/>
        """
        elements.append(Paragraph(metadata_text, self.styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Section titles mapping - combined sections
        section_titles = {
            'company_info': 'Company Information & Metadata',
            'profit_loss': 'Profit & Loss Statement',
            'quarters': 'Quarterly Results',
            'balance_sheet': 'Balance Sheet',
            'cash_flow': 'Cash Flow Statement',
            'price_market': 'Price & Market Data'
        }
        
        # Table of contents
        elements.append(Paragraph("Contents", self.styles['SectionTitle']))
        toc_items = []
        
        for section_key, section_title in section_titles.items():
            if sections.get(section_key):
                toc_items.append(f"• {section_title}")
        
        elements.append(Paragraph('<br/>'.join(toc_items), self.styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Add each section
        for section_key, section_title in section_titles.items():
            if sections.get(section_key) and sections[section_key]:
                print(f"✓ Processing {section_title}: {len(sections[section_key])} rows")
                
                elements.append(PageBreak())
                elements.append(Paragraph(section_title, self.styles['SectionTitle']))
                elements.append(Spacer(1, 0.15*inch))
                
                table = self._create_table(sections[section_key], section_key)
                if table:
                    elements.append(table)
                    elements.append(Spacer(1, 0.2*inch))
        
        # Build PDF
        print(f"\n{'='*80}")
        print("Building PDF document...")
        doc.build(elements)
        
        print(f"{'='*80}")
        print(f"✅ PDF created successfully!")
        print(f"📄 Output file: {self.output_pdf}")
        print(f"📊 File size: {os.path.getsize(self.output_pdf) / 1024:.2f} KB")
        print(f"{'='*80}\n")
        
        self.wb.close()
        return self.output_pdf


def main():
    excel_file = "Apollo Tyres.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Error: File '{excel_file}' not found!")
        return
    
    extractor = DataSheetToPDFExtractor(excel_file)
    extractor.convert()


if __name__ == "__main__":
    main()