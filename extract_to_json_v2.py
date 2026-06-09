from openpyxl import load_workbook
from datetime import datetime
import json
import os
import re


class DataSheetToJSONExtractor:
    def __init__(self, excel_file, output_json=None):
        self.excel_file = excel_file
        self.output_json = output_json or f"{os.path.splitext(excel_file)[0]}_extracted_v2.json"
        self.wb = load_workbook(excel_file, data_only=True)
    
    def _format_date(self, value):
        """Format date value to Mar-17 format"""
        if not value:
            return ''
        
        value_str = str(value).strip()
        
        # If it's already formatted like "Mar-17", keep it
        if re.match(r'[A-Za-z]{3}-\d{2}', value_str):
            return value_str
        
        # If it's a full date string like "2017-03-31 00:00:00"
        if re.match(r'\d{4}-\d{2}-\d{2}', value_str):
            try:
                # Extract year and month
                parts = value_str.split()[0].split('-')
                year = parts[0]
                month = parts[1]
                
                # Convert to Mar-17 format
                month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                month_abbr = month_names[int(month)]
                year_short = year[-2:]
                
                return f"{month_abbr}-{year_short}"
            except:
                pass
        
        # If it's a datetime object
        try:
            if isinstance(value, datetime):
                return value.strftime('%b-%y')
        except:
            pass
        
        return value_str
    
    def _clean_cell_value(self, value):
        """Clean cell value"""
        if value is None:
            return ''
        
        value_str = str(value).strip()
        
        # Remove URLs and screener references
        if 'screener.in' in value_str.lower() or 'http' in value_str.lower():
            return ''
        
        return value_str
    
    def _extract_data_sheet(self):
        """Extract all data from Data Sheet"""
        try:
            ws = self.wb['Data Sheet']
            
            sections = {
                'company_info': [],  # Combined section for company info and metadata
                'profit_loss': [],
                'quarters': [],
                'balance_sheet': [],
                'cash_flow': [],
                'price_market': []  # Combined section
            }
            
            current_section = None
            
            for row in ws.iter_rows(values_only=True):
                # Convert to list and clean values
                row_list = [self._clean_cell_value(cell) for cell in row]
                
                # Skip completely empty rows
                if not any(cell for cell in row_list):
                    continue
                
                first_cell = row_list[0].strip().upper() if row_list[0] else ''
                
                # Identify sections - combine COMPANY NAME and META into company_info
                if 'COMPANY NAME' in first_cell or first_cell == 'META':
                    current_section = 'company_info'
                    if 'COMPANY NAME' in first_cell:
                        sections[current_section].append(row_list)
                    continue
                elif first_cell == 'PROFIT & LOSS':
                    current_section = 'profit_loss'
                    continue
                elif first_cell == 'QUARTERS':
                    current_section = 'quarters'
                    continue
                elif first_cell == 'BALANCE SHEET':
                    current_section = 'balance_sheet'
                    continue
                elif first_cell == 'CASH FLOW:':
                    current_section = 'cash_flow'
                    continue
                elif first_cell == 'PRICE:' or first_cell == 'DERIVED:':
                    # Combine PRICE and DERIVED into price_market section
                    current_section = 'price_market'
                    if first_cell == 'PRICE:':
                        sections[current_section].append(row_list)
                    continue
                elif 'REPORT DATE' in first_cell:
                    # Format report dates properly
                    formatted_dates = [self._format_date(cell) for cell in row_list[1:] if cell]
                    if current_section:
                        sections[current_section].append(['Report Date'] + formatted_dates)
                elif current_section and row_list[0]:
                    # Skip excluded rows
                    if any(x in first_cell for x in ['LATEST VERSION', 'CURRENT VERSION', 'PLEASE DO NOT']):
                        continue
                    sections[current_section].append(row_list)
            
            return sections
            
        except Exception as e:
            print(f"Error extracting data: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
    def _convert_section_to_dict(self, section_data):
        """Convert section data to structured dictionary"""
        if not section_data:
            return None
        
        # Filter out empty rows and prepare data
        filtered_data = []
        
        for row in section_data:
            # Find the last non-empty cell
            last_idx = len(row) - 1
            while last_idx > 0 and not str(row[last_idx]).strip():
                last_idx -= 1
            
            # Keep row up to last non-empty cell
            filtered_row = row[:last_idx + 1]
            
            if filtered_row and any(str(cell).strip() for cell in filtered_row):
                filtered_data.append(filtered_row)
        
        if not filtered_data:
            return None
        
        # Check if first row is Report Date
        first_row_text = ' '.join(str(cell) for cell in filtered_data[0]).upper()
        has_report_date = 'REPORT DATE' in first_row_text
        
        result = {}
        
        if has_report_date and len(filtered_data) > 1:
            # Extract report dates (column headers)
            report_dates = [str(cell).strip() for cell in filtered_data[0][1:] if str(cell).strip()]
            result['report_dates'] = report_dates
            
            # Process data rows
            result['data'] = []
            for row in filtered_data[1:]:
                if not any(str(cell).strip() for cell in row):
                    continue
                
                # Skip if metric name is empty
                if not str(row[0]).strip():
                    continue
                
                # Filter out empty values
                values = [str(cell).strip() for cell in row[1:] if str(cell).strip()]
                
                row_dict = {
                    'metric': str(row[0]).strip(),
                    'values': values
                }
                
                # Only add if there are values
                if row_dict['values']:
                    result['data'].append(row_dict)
        else:
            # No report date row (like company_info)
            result['data'] = []
            for row in filtered_data:
                if not any(str(cell).strip() for cell in row):
                    continue
                
                # Skip if metric name is empty
                if not str(row[0]).strip():
                    continue
                
                # Filter out empty values
                values = [str(cell).strip() for cell in row[1:] if str(cell).strip()]
                
                row_dict = {
                    'metric': str(row[0]).strip(),
                    'values': values
                }
                
                # Only add if there are values
                if row_dict['values']:
                    result['data'].append(row_dict)
        
        return result
    
    def convert(self):
        """Main conversion method"""
        print(f"\n{'='*80}")
        print(f"Extracting data from Data Sheet to JSON (v2)...")
        print(f"{'='*80}\n")
        
        sections = self._extract_data_sheet()
        
        if not sections:
            print("❌ No data extracted!")
            return None
        
        # Extract company name
        company_name = "APOLLO TYRES LTD"
        if sections.get('company_info'):
            for row in sections['company_info']:
                if row[0] and 'COMPANY NAME' in str(row[0]).upper():
                    company_name = row[1] if len(row) > 1 and row[1] else company_name
                    break
        
        json_data = {
            'company': company_name,
            'extraction_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'source_file': self.excel_file,
            'sections': {}
        }
        
        # Section titles mapping - combined sections
        section_titles = {
            'company_info': 'Company Information & Metadata',
            'profit_loss': 'Profit & Loss Statement',
            'quarters': 'Quarterly Results',
            'balance_sheet': 'Balance Sheet',
            'cash_flow': 'Cash Flow Statement',
            'price_market': 'Price & Market Data'
        }
        
        for section_key, section_title in section_titles.items():
            if sections.get(section_key) and sections[section_key]:
                print(f"✓ Processing {section_title}: {len(sections[section_key])} rows")
                
                section_dict = self._convert_section_to_dict(sections[section_key])
                if section_dict and section_dict.get('data'):
                    json_data['sections'][section_key] = {
                        'title': section_title,
                        **section_dict
                    }
        
        print(f"\n{'='*80}")
        print("Writing JSON file...")
        
        with open(self.output_json, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)
        
        print(f"{'='*80}")
        print(f"✅ JSON created successfully!")
        print(f"📄 Output file: {self.output_json}")
        print(f"📊 File size: {os.path.getsize(self.output_json) / 1024:.2f} KB")
        print(f"{'='*80}\n")
        
        self.wb.close()
        return self.output_json


def main():
    excel_file = "Apollo Tyres.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Error: File '{excel_file}' not found!")
        return
    
    extractor = DataSheetToJSONExtractor(excel_file)
    extractor.convert()


if __name__ == "__main__":
    main()